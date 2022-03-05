#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan 19 16:10:42 2021

@author: Nicolas Pourcelot

Lorsqu'on veut fusionner des scores de sources différentes, les formats des names
ne sont pas forcément exactement les mêmes, ce qui complique la fusion.
L'idée est de proposer une fusion raisonnable, qui sera validée ou non ensuite.

"""
from typing import List, Dict, Set, Tuple, NewType

from openpyxl.styles import colors, fills, fonts  # type: ignore
from openpyxl import load_workbook, Workbook  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from fire import Fire  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

CONVERSION = (
    "éèêë" "àâ" "ôö" "ùûü" "îï" "ç" "ñ" "-_",
    "eeee" "aa" "oo" "uuu" "ii" "c" "n" "  ",
)
TABLE = str.maketrans(*CONVERSION)


Name = NewType("Name", str)
Score = NewType("Score", float)
Scores = Tuple[Score, ...]
ScoresForStudent = Dict[Name, Scores]


class ProtectedDict(dict):
    """Only authorize setting keys once. Keys are predefined first."""

    def __setitem__(self, key, value):
        try:
            if self[key] is not None:
                raise ValueError("Key already set !")
            dict.__setitem__(self, key, value)
        except KeyError:
            raise KeyError(f"Unknown key: {key} !")


def norm(name: str) -> Set[str]:
    name = name.casefold()
    # Suppression des accents
    name = name.translate(TABLE)
    return set(name.split())


def match(name1: str, name2: str) -> bool:
    return norm(name1) == norm(name2)


def contain(name1: str, name2: str) -> bool:
    s1 = norm(name1)
    s2 = norm(name2)
    return s1.issubset(s2) or s2.issubset(s1)


def partial_match(name1: str, name2: str) -> bool:
    """Il suffit d'un prénom commun pour que la comparaison réussisse...

    À utiliser en dernier recours, mais peut être utile (étudiant enregistré
    avec le name de famille du père ou de la mère selon la BDD par exemple)."""
    return not norm(name1).isdisjoint(norm(name2))


class FusionData:
    def __init__(self, names):
        self.names: Tuple[Name] = tuple(names)
        self.imported: List = []  # liste de dictionnaires
        self.not_imported: List[ScoresForStudent] = []
        self.number_of_scores_columns_per_sheet: List[int] = []

    def load_sheet_data(
        self, sheet: Worksheet, names: List[Name], name_has_2_cols: bool, n_rows: int
    ) -> ScoresForStudent:
        scores_for_all_students: List[Scores] = []
        for j, col in enumerate(sheet.iter_cols(), start=1):
            if j == 1 or (j == 2 and name_has_2_cols):
                # Skip names columns
                continue

            # Collect scores in the current column.
            vals = tuple(Score(cell.value) for cell in col[:n_rows])
            # If a column is empty, discard all following columns.
            if all(val is None for val in vals):
                break
            scores_for_all_students.append(vals)

        print(" -", len(scores_for_all_students), "column(s) of scores")

        self.number_of_scores_columns_per_sheet.append(len(scores_for_all_students))
        students_scores: ScoresForStudent = {}
        for student, student_scores in zip(names, zip(*scores_for_all_students)):
            students_scores[student] = student_scores
        return students_scores

    def process_sheet_data(self, students_scores: ScoresForStudent) -> None:
        """Associe chaque note à un nom."""

        # Make copies...
        remaining: ScoresForStudent = dict(students_scores)
        names: Set[Name] = set(self.names)
        found: Dict[str, Tuple[str, Scores, int]] = ProtectedDict(dict.fromkeys(names))

        # On essaie déjà de récupérer les noms tels quels.
        used: Set[Name] = set()
        for name in names:
            if name in remaining:
                found[name] = (name, remaining[name], 0)
                used.add(name)
        # Mise à jour
        names -= used
        for name in used:
            remaining.pop(name)

        # On regarde si les noms correspondent en enlevant les accents et en ne tenant
        # pas compte de l'ordre nom/prénom.
        # Puis comparaisons de moins en moins précises...
        for step, comp in enumerate((match, contain, partial_match), start=1):
            used_names = set()
            used_candidates = set()
            for name in names:
                for candidate, scores in remaining.items():
                    if comp(name, candidate):
                        found[name] = (candidate, scores, step)
                        used_names.add(name)
                        used_candidates.add(candidate)
                        # Don't break here !!
                        # We have to be sure there is only one name matching
                        # TODO: deals with the case where several names are matching
            names -= used_names
            for candidate in used_candidates:
                remaining.pop(candidate)

        self.imported.append(found)
        self.not_imported.append(remaining)


def detect_number_of_lines(sheet: Worksheet, column: str = "A") -> int:
    # Detect the table height.
    height = 0
    for height, cell in enumerate(sheet[column], start=1):
        val = cell.value
        if not isinstance(val, str) or val.strip() == "":
            height -= 1
            break
    return height


def get_names(sheet: Worksheet, n_rows: int, name_has_2_cols: bool) -> List[Name]:
    if name_has_2_cols:
        names = [
            Name(f"{a[0].value} {b[0].value}")
            for a, b in zip(sheet[f"A1:A{n_rows}"], sheet[f"B1:B{n_rows}"])
        ]
    else:
        names = [Name(cell[0].value) for cell in sheet[f"A1:A{n_rows}"]]
    return names


def create_new_sheet_with_merge_results(spreadsheet: Workbook, fusion: FusionData) -> None:
    new = spreadsheet.create_sheet("Fusion")

    for i, name in enumerate(sorted(fusion.names), start=1):
        new[f"A{i}"] = name

    # Format for cells that need special attention.
    my_red = colors.Color(rgb="00FF1111")
    my_fill = fills.PatternFill(patternType="solid", fgColor=my_red)

    # Calcul des positions de chaque série de données dans le tableur
    # On mémorise une fois pour toute la position de la 1re colonne correspondant
    # à chaque série de données, pour éviter de la recalculer ensuite.
    positions = []
    pos = 2
    for n in range(len(fusion.imported)):
        positions.append(pos)
        pos += fusion.number_of_scores_columns_per_sheet[n] + 1

    # Fusions réussies
    for i, name in enumerate(sorted(fusion.names), start=1):
        for n, found in enumerate(fusion.imported):
            j = positions[n]
            if found[name] is not None:
                old_name, scores, reliability = found[name]
                new.cell(i, j).value = old_name
                if reliability >= 2:
                    new.cell(i, j).fill = my_fill

                for k, score in enumerate(scores, start=1):
                    new.cell(i, j + k).value = score
                    if reliability >= 2:
                        new.cell(i, j + k).fill = my_fill

    # Éléments pour lesquels la fusion n'a pas fonctionné
    i0 = len(fusion.names) + 2
    all_merged = True
    for n, remaining in enumerate(fusion.not_imported):
        j = positions[n]
        for i, (old_name, scores) in enumerate(sorted(remaining.items()), start=i0):
            all_merged = False
            new.cell(i, j).value = old_name
            new.cell(i, j).fill = my_fill
            for k, score in enumerate(scores, start=1):
                new.cell(i, j + k).value = score
                new.cell(i, j + k).fill = my_fill

    if not all_merged:
        new.cell(i0, 1).value = "Attention, certaines données n'ont pas pu être fusionnées :"
        my_font = fonts.Font(color=my_red, bold=True, italic=True)
        new.cell(i0, 1).font = my_font

    for i, _ in enumerate(new.iter_cols()):
        new.column_dimensions[get_column_letter(i + 1)].width = 25


def fusionner_classeur(filename: str) -> None:
    """Fusionne les données du classeur Excel (document .xlsx).

    La première feuille du tableur doit contenir des noms sur la colonne A.

    Chaque autre feuille doit contenir :
        - des noms sur la colonne A,
        - des notes sur la colonne B.

    Ces colonnes ne doivent pas avoir d'entête.
    Les autres colonnes sont ignorées.
    S'il y a un espace dans la colonne des noms, la lecture des noms s'interrompt.

    Une feuille est générée en fin de document et contient le résultat de la fusion.
    """
    if not filename.endswith(".xlsx"):
        raise RuntimeError(f"File {filename} does not seem to be a .xlsx file.")
    spreadsheet = load_workbook(filename)

    for num, sheet in enumerate(spreadsheet, start=1):
        print(f"Reading {sheet.title!r} sheet...")
        # Guess format: one column for name and surname, or two distinct columns.
        name_has_2_cols = isinstance(sheet["B1"].value, str) and sheet["B1"].value != ""

        # Detect the table height (number of rows).
        n_rows = detect_number_of_lines(sheet, column="A")
        if n_rows == 0:
            # No data in this sheet.
            continue

        print(" -", n_rows, "lines")
        names = get_names(sheet, n_rows, name_has_2_cols)

        if num == 1:
            fusion = FusionData(names)
            # No scores on 1st sheet
        else:
            # noinspection PyUnboundLocalVariable
            students_scores = fusion.load_sheet_data(sheet, names, name_has_2_cols, n_rows)
            fusion.process_sheet_data(students_scores)

    create_new_sheet_with_merge_results(spreadsheet, fusion)
    # Deselect all tabs before activating last one, otherwise a strange bug occurs:
    # the last tab and the previous one are both simultaneously selected when opening
    # the XLSX file in LibreOffice or Excel, which leads to very strange behaviours...
    for sheet in spreadsheet:
        sheet.sheet_view.tabSelected = False
    spreadsheet.active = len(spreadsheet.sheetnames) - 1

    assert "." in filename
    base, ext = filename.split(".")
    spreadsheet.save(f"{base}_output.{ext}")


if __name__ == "__main__":
    Fire(fusionner_classeur)
